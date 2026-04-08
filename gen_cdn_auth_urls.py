#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据 TxCosServiceImpl 的 Type A 鉴权逻辑，从 xlsx 读取 cosKey 列，生成带签名的 CDN 下载链接并输出到新 xlsx。
用法:
  # 单 key 模式：直接传入 cosKey，鉴权 URL 输出到终端
  python gen_cdn_auth_urls.py --key "image/newGoods/xxx.jpeg"

  # xlsx 模式：从 xlsx 读取 cosKey 列，写出到新 xlsx（可不写 output，默认 入参名_COSKEYURL_OUTPUT.xlsx，重复则加数字后缀）
  python gen_cdn_auth_urls.py input.xlsx [output.xlsx] [--key-column COL] [--cdn-domain DOMAIN] [--cdn-key KEY]

  默认 CDN: pharmacytencent.ysbang.cn / 85DC4EUti6u9q14KT1sxEPrldyUkW；可通过环境变量 CDN_DOMAIN、CDN_KEY 覆盖。
  表头：自动查找表头包含 cos_key 或 coskey（不区分大小写）的列。
依赖: pip install openpyxl
"""

import argparse
import hashlib
import os
import sys
from datetime import datetime
from typing import Tuple

def _ensure_openpyxl():
    try:
        from openpyxl import load_workbook, Workbook
        return load_workbook, Workbook
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


# ---------- Java Random 兼容实现（与 TxCosServiceImpl.generateRandom 一致）-----------
MULT = 0x5DEECE66D
ADDEND = 0xB
MASK = (1 << 48) - 1


def _next(seed: int, bits: int) -> Tuple[int, int]:
    seed = (seed * MULT + ADDEND) & MASK
    return seed, (seed >> (48 - bits)) & ((1 << bits) - 1)


def _next_int(seed: int, n: int) -> Tuple[int, int]:
    if n <= 0:
        raise ValueError("n must be positive")
    if n & (n - 1) == 0:
        seed, r = _next(seed, 31)
        return seed, (n * r) >> 31
    while True:
        seed, r = _next(seed, 31)
        if r < 0x7FFFFFFF - (0x7FFFFFFF % n):
            return seed, r % n


def generate_random_java(timestamp: int) -> str:
    """与 Java generateRandom(long timestamp) 行为一致。"""
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"
    # Java: (seed ^ 0x5DEECE66DL) & ((1L << 48) - 1)
    seed = (timestamp ^ MULT) & MASK
    seed, length = _next_int(seed, 100)
    length += 1  # 1-100
    sb = []
    for _ in range(length):
        seed, index = _next_int(seed, len(chars))
        sb.append(chars[index])
    return "".join(sb)


def generate_type_a_auth_url(cos_key: str, cdn_domain: str, cdn_key: str) -> str:
    """与 TxCosServiceImpl.generateTypeAAuthUrl 逻辑一致。"""
    cos_key = (cos_key or "").strip()
    if not cos_key:
        return ""
    url = "https://" + cdn_domain.rstrip("/") + "/" + cos_key.lstrip("/")
    # 当天 0 点时间戳（本地时区）
    now = datetime.now()
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    timestamp = int(midnight.timestamp())
    random_string = generate_random_java(timestamp)
    sign_str = "/" + cos_key + "-" + str(timestamp) + "-" + random_string + "-0-" + cdn_key
    sign_md5 = hashlib.md5(sign_str.encode("utf-8")).hexdigest()
    sign = f"{timestamp}-{random_string}-0-{sign_md5}"
    sep = "&" if "?" in url else "?"
    return url + sep + "sign=" + sign


def main():
    parser = argparse.ArgumentParser(description="从 xlsx 读取 cosKey 或直接传入单个 key，生成 Type A 鉴权 URL")
    parser.add_argument("input_xlsx", nargs="?", default=None, help="输入 xlsx 文件路径（xlsx 模式必填）")
    parser.add_argument("output_xlsx", nargs="?", default=None,
                        help="输出 xlsx 路径；未填时默认 入参名_COSKEYURL_OUTPUT.xlsx，重复则加 _1、_2 等后缀")
    parser.add_argument("--key", "-K", default=None, metavar="COSKEY",
                        help="单 key 模式：直接传入一个 cosKey，将鉴权 URL 输出到终端")
    parser.add_argument("--key-column", "-k", default=None,
                        help="cosKey 所在列名（仅 xlsx 模式）；未指定时自动查找表头包含 cos_key/coskey 的列")
    parser.add_argument("--cdn-domain", default=os.environ.get("CDN_DOMAIN", "pharmacytencent.ysbang.cn"),
                        help="CDN 域名（默认: pharmacytencent.ysbang.cn）")
    parser.add_argument("--cdn-key", default=os.environ.get("CDN_KEY", "85DC4EUti6u9q14KT1sxEPrldyUkW"),
                        help="CDN 鉴权 key（默认已配置）")
    args = parser.parse_args()

    if not args.cdn_domain or not args.cdn_key:
        print("错误: 请提供 --cdn-domain 和 --cdn-key，或设置环境变量 CDN_DOMAIN、CDN_KEY", file=sys.stderr)
        sys.exit(1)

    # 单 key 模式：直接输出 URL 到终端
    if args.key is not None:
        url = generate_type_a_auth_url(args.key.strip(), args.cdn_domain, args.cdn_key)
        print(url)
        return

    # xlsx 模式：至少需要 input_xlsx
    if args.input_xlsx is None:
        print("错误: 请提供 input.xlsx，或使用 --key 传入单个 cosKey", file=sys.stderr)
        sys.exit(1)

    # 未指定输出文件时：默认 入参名_COSKEYURL_OUTPUT.xlsx，重复则加数字后缀
    if args.output_xlsx is None:
        in_dir = os.path.dirname(os.path.abspath(args.input_xlsx))
        in_basename = os.path.splitext(os.path.basename(args.input_xlsx))[0]
        base_path = os.path.join(in_dir, in_basename + "_COSKEYURL_OUTPUT.xlsx")
        args.output_xlsx = base_path
        n = 1
        while os.path.exists(args.output_xlsx):
            args.output_xlsx = os.path.join(in_dir, f"{in_basename}_COSKEYURL_OUTPUT_{n}.xlsx")
            n += 1

    load_workbook, Workbook = _ensure_openpyxl()
    wb = load_workbook(args.input_xlsx, read_only=False, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("输入文件为空或没有数据行。", file=sys.stderr)
        wb.close()
        sys.exit(1)

    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    def header_contains_coskey(col_name: str) -> bool:
        n = (col_name or "").lower().replace(" ", "")
        return "cos_key" in n or "coskey" in n

    key_col_idx = None
    if args.key_column is not None:
        if args.key_column in header:
            key_col_idx = header.index(args.key_column)
    if key_col_idx is None:
        for idx, col_name in enumerate(header):
            if header_contains_coskey(col_name):
                key_col_idx = idx
                break
    if key_col_idx is None:
        print(f"错误: 未找到表头包含 cos_key/coskey 的列，当前表头: {header}", file=sys.stderr)
        wb.close()
        sys.exit(1)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title
    # 新表头：保留原表头 + authUrl
    new_header = header + ["authUrl"]
    out_ws.append(new_header)

    for row in rows[1:]:
        row_list = list(row) if row else []
        while len(row_list) < len(header):
            row_list.append(None)
        cos_key = row_list[key_col_idx]
        cos_key_str = str(cos_key).strip() if cos_key is not None else ""
        if cos_key_str:
            auth_url = generate_type_a_auth_url(cos_key_str, args.cdn_domain, args.cdn_key)
        else:
            auth_url = ""
        out_ws.append(row_list + [auth_url])

    out_wb.save(args.output_xlsx)
    wb.close()
    print(f"已生成 {args.output_xlsx}，共处理 {len(rows) - 1} 行，已保留原列并添加 authUrl 列。")


if __name__ == "__main__":
    main()
